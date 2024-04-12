import ntpath
from openpyxl import load_workbook
from pathlib import Path
from logger import Logger
import settings as st
from core import *
from db import store_securities_to_db

from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String
from sqlalchemy import create_engine
from sqlalchemy.orm import Session


logger = Logger('brokerage_monthly', st.APPLICATION_LOG, write_to_stdout=st.DEBUG_MODE).get()

MAX_EXCEL_ROWS_NUM = 20000
EXCEL_START_COLUMN = 2

STOP_REPORT_STR = 'Владелец: ООО "Компания БКС"'

# 1. Money
MONEY_START_STR = '1. Движение денежных средств'
MONEY_STOP_STR = '2.1. Сделки:'

# 2. Deals
DEALS_START_STR = '2.1. Сделки:'
DEALS_STOP_STR = '3. Активы:'

# 3. Securities
SECURITIES_START_STR = '3. Активы:'
SECURITIES_STOP_STR = '4. Движение Ценных бумаг'
SECURITIES_PORTFOLIO_TOTAL_STR = 'Стоимость портфеля (руб.):'
SECURITIES_PORTFOLIO_TOTAL_BEGIN_COLUMN = 8
SECURITIES_PORTFOLIO_TOTAL_END_COLUMN = 12
SECURITIES_TABLE_START_STR = 'Вид актива'
SECURITIES_TABLE_STOP_STR = 'Итого:'
SECURITIES_TABLE_BEGIN_SUMM_NKD_COLUMN = 9
SECURITIES_TABLE_BEGIN_SUMM_INCLUDING_NKD_COLUMN = 10
SECURITIES_TABLE_END_SUMM_NKD_COLUMN = 13
SECURITIES_TABLE_END_SUMM_INCLUDING_NKD_COLUMN = 14
# SECURITIES_TABLE_COLUMNS = 14

# 4. Securities Transactions
SECURITIES_TRANSACTIONS_START_STR = '4. Движение Ценных бумаг'
SECURITIES_TRANSACTIONS_STOP_STR = '(1*) - оплата проводится со своего (клиента) счета'


class Money:
    def __init__(self, sheet):
        self.sheet = sheet


class Deals:
    def __init__(self, sheet):
        self.sheet = sheet


class Security:
    def __init__(self, secid: str, isin: str | None, sectype: str | None,
                 quantity_begin: float | None, closeprice_begin: float | None, summ_nkd_begin: float | None,
                 summ_including_nkd_begin: float | None, quantity_end: float | None, closeprice_end: float | None,
                 summ_nkd_end: float | None, summ_including_nkd_end: float | None, bidding_organizer: str | None,
                 store_place: str | None, emitent: str | None):

        self.class_name = self.__class__.__name__

        self.secid: str = secid
        self.isin: str | None = isin
        self.sectype: str | None = sectype
        self.quantity_begin: float | None = quantity_begin
        self.closeprice_begin: float | None = closeprice_begin
        self.summ_nkd_begin: float | None = summ_nkd_begin
        self.summ_including_nkd_begin: float | None = summ_including_nkd_begin

        self.quantity_end: float | None = quantity_end
        self.closeprice_end: float | None = closeprice_end
        self.summ_nkd_end: float | None = summ_nkd_end
        self.summ_including_nkd_end: float | None = summ_including_nkd_end

        self.bidding_organizer: str | None = bidding_organizer
        self.store_place: str | None = store_place
        self.emitent: str | None = emitent


class Securities:
    securities: list[Security] = list()

    def __init__(self, sheet, year: int, month: int):
        self.class_name = self.__class__.__name__
        self.sheet = sheet
        self.year = year
        self.month = month
        self.start_column = EXCEL_START_COLUMN
        self.start_row: int | None = None
        self.stop_row: int | None = None
        self.portfolio_total_row: int | None = None
        self.portfolio_total_value_begin_rub: float = 0
        self.portfolio_total_value_end_rub: float = 0
        self.table_start_row: int | None = None
        self.table_stop_row: int | None = None
        self.table_total_row: int | None = None

        self.table_summ_nkd_begin: float = 0
        self.table_summ_including_nkd_begin: float = 0
        self.table_summ_nkd_end: float = 0
        self.table_summ_including_nkd_end: float = 0

        self._find_boundaries()
        self._find_portfolio_total_row()
        self._extract_total_portfolio_values_rub()
        self._find_table_boundaries()
        self._extract_table_summ_values()
        self._load_securities_table()
        self._check_all_securities_summs(6)

    def _find_boundaries(self):
        for i in range(1, MAX_EXCEL_ROWS_NUM+1):
            cell = self.sheet.cell(row=i, column=self.start_column)
            if cell.value == SECURITIES_START_STR:
                self.start_row = cell.row
                break

        if self.start_row:
            for i in range(self.start_row, MAX_EXCEL_ROWS_NUM+1):
                cell = self.sheet.cell(row=i, column=self.start_column)
                if cell.value == SECURITIES_STOP_STR:
                    self.stop_row = cell.row - 1
                    break
        else:
            logger.error(f"{self.class_name}._find_boundaries(): "
                         f"Could not find a Securities start row index")

            raise Exception('Could not find a Securities start row index')

        if not self.stop_row:
            logger.error(f"{self.class_name}._find_boundaries(): "
                         f"Could not find a Securities stop row index")

            raise Exception('Could not find a Securities stop row index')

        logger.info(f"{self.class_name}._find_boundaries(): "
                    f'Securities boundaries found: {self.start_row}, {self.stop_row}')

    def _find_portfolio_total_row(self):
        if self.start_row and self.stop_row:
            for i in range(self.start_row, self.stop_row+1):
                cell = self.sheet.cell(row=i, column=self.start_column)
                if cell.value == SECURITIES_PORTFOLIO_TOTAL_STR:
                    self.portfolio_total_row = cell.row

        else:
            logger.error(f"{self.class_name}._find_portfolio_total_row(): "
                         f"No securities start or/and stop row(s) defined")
            raise Exception('No securities start or/and stop row(s) defined')

        logger.info(f"{self.class_name}._find_portfolio_total_row(): "
                    f'Securities portfolio total row found: {self.portfolio_total_row}')

    def _extract_total_portfolio_values_rub(self):
        if self.portfolio_total_row:
            cell_begin = self.sheet.cell(row=self.portfolio_total_row,
                                         column=SECURITIES_PORTFOLIO_TOTAL_BEGIN_COLUMN)
            self.portfolio_total_value_begin_rub = float(ifnull(cell_begin.value, 0))
            logger.info(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                        f"portfolio_total_value_begin_rub = {self.portfolio_total_value_begin_rub}")

            cell_end = self.sheet.cell(row=self.portfolio_total_row,
                                       column=SECURITIES_PORTFOLIO_TOTAL_END_COLUMN)
            self.portfolio_total_value_end_rub = float(ifnull(cell_end.value, 0))
            logger.info(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                        f"portfolio_total_value_end_rub = {self.portfolio_total_value_end_rub}")
        else:
            logger.error(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                         f"No securities portfolio total row defined")
            raise Exception('No securities portfolio total row defined')

    def _find_table_boundaries(self):
        if self.start_row and self.stop_row:
            for i in range(self.start_row, self.stop_row+1):
                cell = self.sheet.cell(row=i, column=self.start_column)

                if cell.value == SECURITIES_TABLE_START_STR:
                    self.table_start_row = cell.row + 1

            for i in range(self.start_row, self.stop_row+1):
                cell = self.sheet.cell(row=i, column=self.start_column)

                if cell.value == SECURITIES_TABLE_STOP_STR:
                    self.table_total_row = cell.row
                    self.table_stop_row = cell.row - 1

        else:
            logger.error(f"{self.class_name}._find_table_boundaries(): "
                         f"No securities start or/and stop row(s) defined")
            raise Exception('No securities start or/and stop row(s) defined')

        logger.info(f"{self.class_name}._find_table_boundaries(): "
                    f'Securities table boundaries found: {self.table_start_row}, {self.table_stop_row}')

    def _extract_table_summ_values(self):
        if self.table_total_row:
            cell_begin_summ_nkd = self.sheet.cell(row=self.table_total_row,
                                                  column=SECURITIES_TABLE_BEGIN_SUMM_NKD_COLUMN)
            cell_begin_summ_including_nkd = self.sheet.cell(row=self.table_total_row,
                                                            column=SECURITIES_TABLE_BEGIN_SUMM_INCLUDING_NKD_COLUMN)
            cell_end_summ_nkd = self.sheet.cell(row=self.table_total_row,
                                                column=SECURITIES_TABLE_END_SUMM_NKD_COLUMN)
            cell_end_summ_including_nkd = self.sheet.cell(row=self.table_total_row,
                                                          column=SECURITIES_TABLE_END_SUMM_INCLUDING_NKD_COLUMN)

            self.table_summ_nkd_begin = float(ifnull(cell_begin_summ_nkd.value, 0))
            self.table_summ_including_nkd_begin = float(ifnull(cell_begin_summ_including_nkd.value, 0))
            self.table_summ_nkd_end = float(ifnull(cell_end_summ_nkd.value, 0))
            self.table_summ_including_nkd_end = float(ifnull(cell_end_summ_including_nkd.value, 0))

            logger.info(f"{self.class_name}._extract_table_summ_values(): "
                        f'Securities table summ values found: {self.table_summ_nkd_begin}, '
                        f'{self.table_summ_including_nkd_begin}, {self.table_summ_nkd_end}, '
                        f'{self.table_summ_including_nkd_end}')
        else:
            logger.error(f"{self.class_name}._extract_table_summ_values(): "
                         f"No securities table total row defined")
            raise Exception('No securities table total row defined')

    def _load_securities_table(self):
        if self.table_start_row and self.table_stop_row:
            self.securities.clear()
            for i in range(self.table_start_row, self.table_stop_row+1):
                cell = self.sheet.cell(row=i, column=2)
                if not cell.value:
                    raise Exception('secid value must not be None')
                secid: str = cell.value

                cell = self.sheet.cell(row=i, column=4)
                isin: str | None = cell.value

                cell = self.sheet.cell(row=i, column=6)
                sectype: str | None = cell.value

                cell = self.sheet.cell(row=i, column=7)
                quantity_begin: float | None = cell.value

                cell = self.sheet.cell(row=i, column=8)
                closeprice_begin: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=9)
                summ_nkd_begin: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=10)
                summ_including_nkd_begin: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=11)
                quantity_end: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=12)
                closeprice_end: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=13)
                summ_nkd_end: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=14)
                summ_including_nkd_end: float | None = tofloat(cell.value)

                cell = self.sheet.cell(row=i, column=15)
                bidding_organizer: str | None = cell.value

                cell = self.sheet.cell(row=i, column=16)
                store_place: str | None = cell.value

                cell = self.sheet.cell(row=i, column=17)
                emitent: str | None = cell.value

                security = Security(secid, isin, sectype, quantity_begin, closeprice_begin, summ_nkd_begin,
                                    summ_including_nkd_begin, quantity_end, closeprice_end, summ_nkd_end,
                                    summ_including_nkd_end, bidding_organizer, store_place, emitent)

                self.securities.append(security)

            logger.info(f"{self.class_name}._load_securities_table(): "
                        f"{len(self.securities)} Securities loaded")

        else:
            logger.error(f"{self.class_name}._load_securities_table(): "
                         f"No securities start or/and stop row(s) defined")
            raise Exception('No securities start or/and stop row(s) defined')

        logger.info(f"{self.class_name}._load_securities_table(): "
                    f'Securities table successfully loaded')

    def _check_all_securities_summs(self, precision):
        if (self.portfolio_total_value_begin_rub == self.table_summ_including_nkd_begin) and \
           (self.portfolio_total_value_end_rub == self.table_summ_including_nkd_end):
            logger.info(f"{self.class_name}._check_all_securities_summs(): "
                        f"Summs in total portfolio table correspond total summs in a detailed table")

        else:
            logger.error(f"{self.class_name}._check_all_securities_summs(): "
                         f"Summs in total portfolio table do not mach total summs in a detailed table")
            raise Exception('Summs in total portfolio table do not mach total summs in a detailed table')

        table_summ_nkd_begin: float = 0
        table_summ_including_nkd_begin: float = 0
        table_summ_nkd_end: float = 0
        table_summ_including_nkd_end: float = 0

        for sec in self.securities:
            if sec.summ_nkd_begin:
                table_summ_nkd_begin += sec.summ_nkd_begin
            if sec.summ_including_nkd_begin:
                table_summ_including_nkd_begin += sec.summ_including_nkd_begin
            if sec.summ_nkd_end:
                table_summ_nkd_end += sec.summ_nkd_end
            if sec.summ_including_nkd_end:
                table_summ_including_nkd_end += sec.summ_including_nkd_end

        logger.info(f"{self.class_name}._check_all_securities_summs(): "
                    f"Calculated summs: {table_summ_nkd_begin}, {table_summ_including_nkd_begin}, "
                    f"{table_summ_nkd_end}, {table_summ_including_nkd_end}")

        if round(table_summ_nkd_begin, precision) == round(self.table_summ_nkd_begin, precision) and \
           round(table_summ_including_nkd_begin, precision) == round(self.table_summ_including_nkd_begin, precision) and \
           round(table_summ_nkd_end, precision) == round(self.table_summ_nkd_end, precision) and \
           round(table_summ_including_nkd_end, precision) == round(self.table_summ_including_nkd_end, precision):
            logger.info(f"{self.class_name}._check_all_securities_summs(): "
                        f"Summs in total row correspond summs of all rows")
        else:
            logger.error(f"{self.class_name}._check_all_securities_summs(): "
                         f"Summs in total row do not mach summs of all rows")
            raise Exception("Summs in total row do not mach summs of all rows")

    def _store_to_db(self):
        store_securities_to_db()  # (self.securities, self.year, self.month)


class SecuritiesTransactions:
    def __init__(self, sheet):
        self.sheet = sheet


class BrokerMonthly:
    # 1. Money
    money: Money
    # 2. Deals
    deals: Deals
    # 3. Securities
    securities: Securities
    # 4. Securities Transactions
    securities_transactions: SecuritiesTransactions

    def __init__(self, report_path):
        self.year: int = 0
        self.month: int = 0
        self.stop_report_str = STOP_REPORT_STR
        self.class_name = self.__class__.__name__
        self.report_path: Path = report_path
        self.workbook = load_workbook(self.report_path)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]

        self._extract_year_and_month()

        self.money = Money(self.sheet)
        self.deals = Deals(self.sheet)
        self.securities = Securities(self.sheet, self.year, self.month)
        self.securities_transactions = SecuritiesTransactions(self.sheet)

    def _extract_year_and_month(self):
        bare_file_name = get_file_name(ntpath.basename(self.report_path))
        str_year_month = bare_file_name.split('_ALL_')[-1]
        self.year = 2000 + int(str_year_month.split('-')[0])
        self.month = int(str_year_month.split('-')[-1])
        logger.info(f"Year {self.year} and Month {self.month} were extracted")