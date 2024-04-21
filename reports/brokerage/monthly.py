import ntpath

from openpyxl import load_workbook
from logger import Logger
import settings as st
from core import *

from sqlalchemy import ForeignKey
from sqlalchemy.orm import relationship
from sqlalchemy import Column, Integer, String, DateTime, Boolean
from sqlalchemy.sql import func
from db import Base

from reports.brokerage.securities import BrokerageMonthlySecurities
from reports.brokerage.money import BrokerageMonthlyMoney
from reports.brokerage.deals import BrokerageMonthlyDeals

MAX_EXCEL_ROWS_NUM = 20000
EXCEL_START_COLUMN = 2

SCHEMA_NAME = 'reports'
STOP_REPORT_STR = '(1*) - оплата проводится со своего (клиента) счета'

BROKERAGE_MONTHLY_MONEY_MARKER = '1. Движение денежных средств'
BROKERAGE_MONTHLY_MONEY_MARKER1 = 'Итого по валюте Рубль:'
# BROKERAGE_MONTHLY_MONEY_MARKER2 = '"-" - задолженность клиента перед компанией(17*)'
BROKERAGE_MONTHLY_DEALS_MARKER = '2.1. Сделки:'
BROKERAGE_MONTHLY_SECURITIES_MARKER = '3. Активы:'
BROKERAGE_MONTHLY_SECURITIES_TRANSACTIONS_MARKER = '4. Движение Ценных бумаг'

logger = Logger('brokerage_monthly', st.APPLICATION_LOG, write_to_stdout=st.DEBUG_MODE).get()


class BrokerageMonthly(Base):
    __tablename__ = 'brokerage_monthly'
    __table_args__ = {"schema": SCHEMA_NAME}

    id = Column(Integer, primary_key=True)
    money = relationship("BrokerageMonthlyMoney", uselist=False, backref="report")
    # deals
    securities = relationship("BrokerageMonthlySecurities", uselist=False, backref="report")
    # securities_transactions

    report_path = Column(String, nullable=False)
    year = Column(Integer, nullable=False)
    month = Column(Integer, nullable=False)
    stop_report_str = Column(String, nullable=False)

    chapter_money_operations = Column(Boolean, nullable=False)
    chapter_money_fees = Column(Boolean, nullable=False)
    chapter_deals = Column(Boolean, nullable=False)

    inserted = Column(DateTime(), server_default=func.now())  # timezone=True
    updated = Column(DateTime(), onupdate=func.now())  # timezone=True

    def __init__(self, report_path):
        self.start_column = EXCEL_START_COLUMN
        self.year: int = 0
        self.month: int = 0
        self.stop_report_str = STOP_REPORT_STR
        self.class_name = self.__class__.__name__
        self.report_path: str = str(report_path)
        self.workbook = load_workbook(self.report_path)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]

        self.chapter_money_start_pos: int = 0
        self.chapter_money_stop_pos: int = 0
        self.chapter_money_operations: bool = False
        self.chapter_money_fees: bool = False

        self.chapter_deals: bool = False
        self.chapter_deals_start_pos: int = 0
        self.chapter_deals_stop_pos: int = 0

        self.chapter_securities_start_pos: int = 0
        self.chapter_securities_stop_pos: int = 0

        self.chapter_securities_transactions_start_pos: int = 0
        self.chapter_securities_transactions_stop_pos: int = 0

        self._extract_year_and_month()

        self._detect_chapters()

        self.money = BrokerageMonthlyMoney(self.sheet,
                                           self.chapter_money_start_pos,
                                           self.chapter_money_stop_pos,
                                           self.chapter_money_operations,
                                           self.chapter_money_fees)

        if self.chapter_deals:
            self.deals = BrokerageMonthlyDeals(self.sheet,
                                               self.chapter_deals_start_pos,
                                               self.chapter_deals_stop_pos)

        self.securities = BrokerageMonthlySecurities(self.sheet,
                                                     self.chapter_securities_start_pos,
                                                     self.chapter_securities_stop_pos)

        # self.securities_transactions = SecuritiesTransactions(self.sheet)

    def __repr__(self):
        return f"<BrokerageMonthly({self.year}, {self.month})>"

    def _extract_year_and_month(self):
        bare_file_name = get_file_name(ntpath.basename(self.report_path))
        str_year_month = bare_file_name.split('_ALL_')[-1]
        self.year = 2000 + int(str_year_month.split('-')[0])
        self.month = int(str_year_month.split('-')[-1])
        logger.info(f"Year {self.year} and Month {self.month} were extracted")

    def _detect_chapters(self):

        money_marker1_occurrences = 0

        for i in range(1, MAX_EXCEL_ROWS_NUM + 1):
            cell = self.sheet.cell(row=i, column=self.start_column)

            if cell.value == BROKERAGE_MONTHLY_MONEY_MARKER1:
                money_marker1_occurrences += 1

                # if self.chapter_money_stop_pos == 0:
                self.chapter_money_stop_pos = cell.row

            if cell.value == BROKERAGE_MONTHLY_MONEY_MARKER:
                self.chapter_money_start_pos = cell.row

            if cell.value == BROKERAGE_MONTHLY_DEALS_MARKER:
                self.chapter_deals = True
                self.chapter_deals_start_pos = cell.row

                if self.chapter_money_stop_pos == 0:
                    self.chapter_money_stop_pos = cell.row - 1

                logger.info(f"{self.class_name}._detect_chapters(): "
                            f"chapter_deals was found")

            if cell.value == BROKERAGE_MONTHLY_SECURITIES_MARKER:
                self.chapter_securities_start_pos = cell.row

                if self.chapter_deals:
                    self.chapter_deals_stop_pos = cell.row - 1

                if self.chapter_money_stop_pos == 0:
                    self.chapter_money_stop_pos = cell.row - 1

            if cell.value == BROKERAGE_MONTHLY_SECURITIES_TRANSACTIONS_MARKER:
                self.chapter_securities_stop_pos = cell.row - 1
                self.chapter_securities_transactions_start_pos = cell.row

            if cell.value == STOP_REPORT_STR:
                self.chapter_securities_transactions_stop_pos = cell.row - 1

        if money_marker1_occurrences == 0:
            self.chapter_money_operations = False
            self.chapter_money_fees = False
            logger.info(f"{self.class_name}._detect_chapters(): "
                        f"None of the chapter money operations or chapter money fees was found")
        elif money_marker1_occurrences == 1:
            self.chapter_money_operations = True
            self.chapter_money_fees = False
            logger.info(f"{self.class_name}._detect_chapters(): "
                        f"chapter money operations was found but chapter money fees was not found")
        elif money_marker1_occurrences == 2:
            self.chapter_money_operations = True
            self.chapter_money_fees = True
            logger.info(f"{self.class_name}._detect_chapters(): "
                        f"chapter money operations and chapter money fees were found")
        else:
            logger.error(f"{self.class_name}._detect_chapters(): "
                         f"Error in business logic detecting report chapters")
            raise Exception('Error in business logic detecting report chapters')
