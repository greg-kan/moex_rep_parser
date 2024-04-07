from openpyxl import load_workbook
from pathlib import Path
from logger import Logger
import settings as st

logger = Logger('brokerage_monthly', st.APPLICATION_LOG, write_to_stdout=st.DEBUG_MODE).get()

MAX_EXCEL_ROWS_NUM = 20000
EXCEL_START_COLUMN = 2

STOP_REPORT_STR = 'Владелец: ООО "Компания БКС"'

# 1. Money
MONEY_START_STR = '1. Движение денежных средств'
MONEY_STOP_STR = '2.1. Сделки:'

# 2. Deals
DEALS_START_STR = '2.1. Сделки:'
DEALS_STOP_STR = '3. Активы:'

# 3. Securities
SECURITIES_START_STR = '3. Активы:'
SECURITIES_STOP_STR = '4. Движение Ценных бумаг'
SECURITIES_PORTFOLIO_TOTAL_STR = 'Стоимость портфеля (руб.):'
SECURITIES_PORTFOLIO_TOTAL_BEGIN_COLUMN = 8
SECURITIES_PORTFOLIO_TOTAL_END_COLUMN = 12
SECURITIES_TABLE_START_STR = 'Вид актива'
SECURITIES_TABLE_STOP_STR = 'Итого:'
SECURITIES_TABLE_BEGIN_SUMM_NKD_COLUMN = 9
SECURITIES_TABLE_BEGIN_SUMM_INCLUDING_NKD_COLUMN = 10
SECURITIES_TABLE_END_SUMM_NKD_COLUMN = 13
SECURITIES_TABLE_END_SUMM_INCLUDING_NKD_COLUMN = 14

# 4. Securities Transactions
SECURITIES_TRANSACTIONS_START_STR = '4. Движение Ценных бумаг'
SECURITIES_TRANSACTIONS_STOP_STR = '(1*) - оплата проводится со своего (клиента) счета'


class Money:
    def __init__(self, sheet):
        self.sheet = sheet


class Deals:
    def __init__(self, sheet):
        self.sheet = sheet


class Security:
    def __init__(self):
        self.class_name = self.__class__.__name__


class Securities:
    securities: list[Security] = list()

    def __init__(self, sheet):
        self.class_name = self.__class__.__name__
        self.sheet = sheet
        self.start_column = EXCEL_START_COLUMN
        self.securities: list[dict] | None = None
        self.start_row: int | None = None
        self.stop_row: int | None = None
        self.portfolio_total_row: int | None = None
        self.portfolio_total_value_begin_rub: float = 0
        self.portfolio_total_value_end_rub: float = 0
        self.table_start_row: int | None = None
        self.table_stop_row: int | None = None
        self.table_total_row: int | None = None

        self.table_begin_summ_nkd: float = 0
        self.table_begin_summ_including_nkd: float = 0
        self.table_end_summ_nkd: float = 0
        self.table_end_summ_including_nkd: float = 0

        self._find_boundaries()
        self._find_portfolio_total_row()
        self._extract_total_portfolio_values_rub()
        self._find_table_boundaries()
        self._extract_table_summ_values()
        self._check_table_total_summs()

        self.load()

    def _find_boundaries(self):
        for i in range(1, MAX_EXCEL_ROWS_NUM):
            cell = self.sheet.cell(row=i, column=self.start_column)
            if cell.value == SECURITIES_START_STR:
                self.start_row = cell.row
                break

        if self.start_row:
            for i in range(self.start_row, MAX_EXCEL_ROWS_NUM):
                cell = self.sheet.cell(row=i, column=self.start_column)
                if cell.value == SECURITIES_STOP_STR:
                    self.stop_row = cell.row - 1
                    break
        else:
            logger.error(f"{self.class_name}._find_boundaries(): "
                         f"Could not find a Securities start row index")

            raise Exception('Could not find a Securities start row index')

        if not self.stop_row:
            logger.error(f"{self.class_name}._find_boundaries(): "
                         f"Could not find a Securities stop row index")

            raise Exception('Could not find a Securities stop row index')

        logger.info(f"{self.class_name}._find_boundaries(): "
                    f'Securities boundaries found: {self.start_row}, {self.stop_row}')

    def _find_portfolio_total_row(self):
        if self.start_row and self.stop_row:
            for i in range(self.start_row, self.stop_row):
                cell = self.sheet.cell(row=i, column=self.start_column)
                if cell.value == SECURITIES_PORTFOLIO_TOTAL_STR:
                    self.portfolio_total_row = cell.row

        else:
            logger.error(f"{self.class_name}._find_portfolio_total_row(): "
                         f"No securities start or/and stop row(s) defined")
            raise Exception('No securities start or/and stop row(s) defined')

        logger.info(f"{self.class_name}._find_portfolio_total_row(): "
                    f'Securities portfolio total row found: {self.portfolio_total_row}')

    def _extract_total_portfolio_values_rub(self):
        if self.portfolio_total_row:
            cell_begin = self.sheet.cell(row=self.portfolio_total_row,
                                         column=SECURITIES_PORTFOLIO_TOTAL_BEGIN_COLUMN)
            self.portfolio_total_value_begin_rub = float(cell_begin.value)
            logger.info(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                        f"portfolio_total_value_begin_rub = {self.portfolio_total_value_begin_rub}")

            cell_end = self.sheet.cell(row=self.portfolio_total_row,
                                       column=SECURITIES_PORTFOLIO_TOTAL_END_COLUMN)
            self.portfolio_total_value_end_rub = float(cell_end.value)
            logger.info(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                        f"portfolio_total_value_end_rub = {self.portfolio_total_value_end_rub}")
        else:
            logger.error(f"{self.class_name}._extract_total_portfolio_values_rub(): "
                         f"No securities portfolio total row defined")
            raise Exception('No securities portfolio total row defined')

    def _find_table_boundaries(self):
        if self.start_row and self.stop_row:
            for i in range(self.start_row, self.stop_row):
                cell = self.sheet.cell(row=i, column=self.start_column)

                if cell.value == SECURITIES_TABLE_START_STR:
                    self.table_start_row = cell.row + 1

            for i in range(self.start_row, self.stop_row):
                cell = self.sheet.cell(row=i, column=self.start_column)

                if cell.value == SECURITIES_TABLE_STOP_STR:
                    self.table_total_row = cell.row
                    self.table_stop_row = cell.row - 1

        else:
            logger.error(f"{self.class_name}._find_table_boundaries(): "
                         f"No securities start or/and stop row(s) defined")
            raise Exception('No securities start or/and stop row(s) defined')

        logger.info(f"{self.class_name}._find_table_boundaries(): "
                    f'Securities table boundaries found: {self.table_start_row}, {self.table_stop_row}')

    def _extract_table_summ_values(self):
        if self.table_total_row:
            cell_begin_summ_nkd = self.sheet.cell(row=self.table_total_row,
                                                  column=SECURITIES_TABLE_BEGIN_SUMM_NKD_COLUMN)
            cell_begin_summ_including_nkd = self.sheet.cell(row=self.table_total_row,
                                                            column=SECURITIES_TABLE_BEGIN_SUMM_INCLUDING_NKD_COLUMN)
            cell_end_summ_nkd = self.sheet.cell(row=self.table_total_row,
                                                column=SECURITIES_TABLE_END_SUMM_NKD_COLUMN)
            cell_end_summ_including_nkd = self.sheet.cell(row=self.table_total_row,
                                                          column=SECURITIES_TABLE_END_SUMM_INCLUDING_NKD_COLUMN)
            self.table_begin_summ_nkd = float(cell_begin_summ_nkd.value)
            self.table_begin_summ_including_nkd = float(cell_begin_summ_including_nkd.value)
            self.table_end_summ_nkd = float(cell_end_summ_nkd.value)
            self.table_end_summ_including_nkd = float(cell_end_summ_including_nkd.value)

            logger.info(f"{self.class_name}._extract_table_summ_values(): "
                        f'Securities table summ values found: {self.table_begin_summ_nkd}, '
                        f'{self.table_begin_summ_including_nkd}, {self.table_end_summ_nkd}, '
                        f'{self.table_end_summ_including_nkd}')
        else:
            logger.error(f"{self.class_name}._extract_table_summ_values(): "
                         f"No securities table total row defined")
            raise Exception('No securities table total row defined')

    def _check_table_total_summs(self):
        # TODO: After loading securities table add check summ of all rows

        if (self.portfolio_total_value_begin_rub == self.table_begin_summ_including_nkd) and \
           (self.portfolio_total_value_end_rub == self.table_end_summ_including_nkd):
            logger.info(f"{self.class_name}._check_table_total_summs(): "
                        f"Summs in total portfolio table correspond total summs in a detailed table")

        else:
            logger.error(f"{self.class_name}._check_table_total_summs(): "
                         f"Summs in total portfolio table do not mach total summs in a detailed table")
            raise Exception('Summs in total portfolio table do not mach total summs in a detailed table')









    def load(self):
        pass


class SecuritiesTransactions:
    def __init__(self, sheet):
        self.sheet = sheet


class BrokerageMonthly:
    # 1. Money
    money: Money
    # 2. Deals
    deals: Deals
    # 3. Securities
    securities: Securities
    # 4. Securities Transactions
    securities_transactions: SecuritiesTransactions

    def __init__(self, report_path):
        self.stop_report_str = STOP_REPORT_STR
        self.class_name = self.__class__.__name__
        self.report_path: Path = report_path
        self.workbook = load_workbook(self.report_path)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]

        self.money = Money(self.sheet)
        self.deals = Deals(self.sheet)
        self.securities = Securities(self.sheet)
        self.securities_transactions = SecuritiesTransactions(self.sheet)

    def feature_method(self):
        pass


